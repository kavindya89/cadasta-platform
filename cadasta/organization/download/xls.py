import os
from openpyxl import Workbook
from django.conf import settings
from django.contrib.contenttypes.models import ContentType

from .base import Exporter


MIME_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class XLSExporter(Exporter):
    def write_items(self, worksheet, queryset, content_type, model_attrs):
        schema_attrs = self.get_schema_attrs(content_type)

        # write column labels
        worksheet.append(model_attrs + [a.name for a in schema_attrs])

        # write data
        for i, item in enumerate(queryset):
            values = self.get_values(item, model_attrs, schema_attrs)
            worksheet.append(values)

    def write_locations(self):
        worksheet = self.workbook.create_sheet(title='locations')
        locations = self.project.spatial_units.all()
        content_type = ContentType.objects.get(app_label='spatial',
                                               model='spatialunit')
        self.write_items(worksheet, locations, content_type,
                         ['id', 'geometry.ewkt'])

    def write_parties(self):
        worksheet = self.workbook.create_sheet(title='parties')
        parties = self.project.parties.all()
        content_type = ContentType.objects.get(app_label='party',
                                               model='party')
        self.write_items(worksheet, parties, content_type,
                         ['id', 'name', 'type'])

    def write_relationships(self):
        worksheet = self.workbook.create_sheet(title='relationships')
        relationships = self.project.tenure_relationships.all()
        content_type = ContentType.objects.get(app_label='party',
                                               model='tenurerelationship')
        self.write_items(worksheet, relationships, content_type,
                         ['party_id', 'spatial_unit_id', 'tenure_type.label'])

    def make_download(self, f_name):
        path = os.path.join(settings.MEDIA_ROOT, 'temp/{}.xlsx'.format(f_name))
        self.workbook = Workbook()
        self.workbook.remove_sheet(self.workbook['Sheet'])

        self.write_locations()
        self.write_parties()
        self.write_relationships()

        self.workbook.save(filename=path)
        return path, MIME_TYPE
